import json
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def generate_excel(profiles: list[dict]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "LinkedIn Profiles"

    headers = ["Name", "Profile URL", "Bio Snippet", "Matched Keywords"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, profile in enumerate(profiles, 2):
        ws.cell(row=row_idx, column=1, value=profile["name"])
        ws.cell(row=row_idx, column=2, value=profile["profile_url"])
        ws.cell(row=row_idx, column=3, value=profile.get("snippet", ""))

        keywords = profile.get("matched_keywords", "[]")
        if isinstance(keywords, str):
            keywords = json.loads(keywords)
        ws.cell(row=row_idx, column=4, value=", ".join(keywords))

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 30

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
