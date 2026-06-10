import pytest
import json
from io import BytesIO
from openpyxl import load_workbook

from app.services.excel_export import generate_excel


def test_generate_excel_basic():
    profiles = [
        {
            "name": "Alice Smith",
            "profile_url": "https://linkedin.com/in/alice",
            "snippet": "AI expert in healthcare",
            "matched_keywords": json.dumps(["Healthcare AI", "Medtech"]),
        },
        {
            "name": "Bob Jones",
            "profile_url": "https://linkedin.com/in/bob",
            "snippet": "Medical device founder",
            "matched_keywords": json.dumps(["Medtech"]),
        },
    ]
    output = generate_excel(profiles)
    assert isinstance(output, BytesIO)

    wb = load_workbook(output)
    ws = wb.active
    assert ws.title == "LinkedIn Profiles"

    headers = [ws.cell(row=1, column=c).value for c in range(1, 5)]
    assert headers == ["Name", "Profile URL", "Bio Snippet", "Matched Keywords"]

    assert ws.cell(row=2, column=1).value == "Alice Smith"
    assert ws.cell(row=2, column=2).value == "https://linkedin.com/in/alice"
    assert ws.cell(row=2, column=3).value == "AI expert in healthcare"
    assert ws.cell(row=2, column=4).value == "Healthcare AI, Medtech"

    assert ws.cell(row=3, column=1).value == "Bob Jones"


def test_generate_excel_empty():
    output = generate_excel([])
    wb = load_workbook(output)
    ws = wb.active
    assert ws.max_row == 1
